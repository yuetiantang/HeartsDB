import openpyxl
from datetime import datetime
from datetime import date
import glob

from src.util import *


def verify_tournament_file_integrity(path, mute=False):
    if not mute:
        print(f'Verifying integrity of file \'{path}\'')
    tournament_wb = openpyxl.load_workbook(path)
    game_sheet_number = len(tournament_wb.worksheets)
    for game_sheet_idx, game_sheet in enumerate(tournament_wb.worksheets):
        # check if name appears in system
        player_name_cells = game_sheet['A1':'D1'][0]
        for name_cell in player_name_cells:
            if not verify_player_name(name_cell.value):
                print(f'File \'{path}\', Sheet \'{tournament_wb.sheetnames[game_sheet_idx]}\' '
                      f'presents unresolved player name: \'{name_cell.value}\'')
        # check if round numbers the same
        end_of_game_flag = False
        error = False
        round_number = 0
        for row_idx, round_row in enumerate(game_sheet.iter_rows(min_row=2, max_col=4, max_row=17)):
            for score_cell in round_row:
                if end_of_game_flag and score_cell.value is not None:
                    print(f'File \'{path}\', Sheet \'{tournament_wb.sheetnames[game_sheet_idx]}\', Row \'{row_idx+2}\' '
                          f'presents unmatched round number: \'{score_cell.value}\'')
                    error = True
                    break
                elif not end_of_game_flag and score_cell.value is None:
                    round_number = row_idx
                    end_of_game_flag = True
            if error:
                break
        if error:
            continue
        # check if round score increment valid
        last_round_scores = [0, 0, 0, 0]
        this_round_scores = [0, 0, 0, 0]
        for row_idx, round_row in enumerate(game_sheet.iter_rows(min_row=2, max_col=4, max_row=round_number+1)):
            shoot_the_moon = 0
            round_score_increment = 0
            for player_idx, score_cell in enumerate(round_row):
                if type(score_cell.value) is not int:
                    print(f'File \'{path}\', Sheet \'{tournament_wb.sheetnames[game_sheet_idx]}\', Row \'{row_idx+2}\' '
                          f'presents invalid score type: \'{score_cell.value}\'')
                    break
                this_round_scores[player_idx] = score_cell.value
                if row_idx == 0 and game_sheet_number == game_sheet_idx + 1:
                    last_round_scores[player_idx] = this_round_scores[player_idx]
                    continue
                player_score_increment = this_round_scores[player_idx] - last_round_scores[player_idx]
                if player_score_increment >= 27 or player_score_increment <= -1:
                    print(f'File \'{path}\', Sheet \'{tournament_wb.sheetnames[game_sheet_idx]}\', Row \'{row_idx+2}\' '
                          f'presents invalid score value: \'{score_cell.value}\'')
                    break
                elif player_score_increment == 26:
                    shoot_the_moon += 1
                round_score_increment += player_score_increment
            if row_idx == 0 and game_sheet_number == game_sheet_idx + 1:
                continue
            if shoot_the_moon == 0 and round_score_increment != 26:
                print(f'File \'{path}\', Sheet \'{tournament_wb.sheetnames[game_sheet_idx]}\', Row \'{row_idx+2}\' '
                      f'presents invalid score value, scores add up to: \'{round_score_increment}\' (not shot the moon)')
            elif shoot_the_moon == 3 and round_score_increment != 78:
                print(f'File \'{path}\', Sheet \'{tournament_wb.sheetnames[game_sheet_idx]}\', Row \'{row_idx+2}\' '
                      f'presents invalid score value, scores add up to: \'{round_score_increment}\' (shot the moon)')
            elif (shoot_the_moon == 0 and round_score_increment == 26) or \
                 (shoot_the_moon == 3 and round_score_increment == 78):
                pass
            else:
                print(f'File \'{path}\', Sheet \'{tournament_wb.sheetnames[game_sheet_idx]}\', Row \'{row_idx+2}\' '
                      f'presents invalid score value, scores add up to: \'{round_score_increment}\' (shot the moon?)')
            for player_idx in range(4):
                last_round_scores[player_idx] = this_round_scores[player_idx]
    if not mute:
        print('Done.')
def update_player_stats(skip_keyword=''):
    stat_wb = openpyxl.Workbook()
    stat_ws = stat_wb.active
    stat_ws.title = "player stats"
    stat_ws.cell(row=1, column=1, value='选手')
    stat_ws.cell(row=1, column=2, value='轮次')
    stat_ws.cell(row=1, column=3, value='场次')
    stat_ws.cell(row=1, column=4, value='总分')
    for player_idx, player in enumerate(Players):
        stat_ws.cell(row=player_idx+2, column=1, value=player.value[1])
    # Buffers
    round_number_list = [0 for i in range(PLAYER_COUNT)]
    game_number_list = [0 for i in range(PLAYER_COUNT)]
    total_score_list = [0 for i in range(PLAYER_COUNT)]
    # Read in
    for data_file in glob.glob('../data/*.xlsx'):
        if skip_keyword != '' and data_file.find(skip_keyword) != -1:
            continue
        print(data_file)
        tournament_wb = openpyxl.load_workbook(data_file)
        for game_sheet_idx, game_sheet in enumerate(tournament_wb.worksheets):
            # sum up player scores, rounds, games
            player_name_cells = game_sheet['A1':'D1'][0]
            for column_idx, name_cell in enumerate(player_name_cells):  # a cell may contain more than 1 player
                for player in convert_name_to_player(name_cell.value):
                    game_number_list[player.value[0]-1] += 1  # player_idx starts from 1
                    if game_sheet_idx == len(tournament_wb.sheetnames) - 1:
                        round_number_list[player.value[0]-1] += game_sheet.max_row - 2
                        total_score_list[player.value[0]-1] += game_sheet.cell(row=game_sheet.max_row,
                                                                            column=column_idx+1).value - \
                                                            game_sheet.cell(row=2, column=column_idx+1).value
                    else:
                        round_number_list[player.value[0]-1] += game_sheet.max_row - 1
                        total_score_list[player.value[0]-1] += game_sheet.cell(row=game_sheet.max_row,
                                                                              column=column_idx+1).value
    # Write in
    for row_idx, round_number in enumerate(round_number_list):
        stat_ws.cell(row=row_idx+2, column=2, value=round_number)
    for row_idx, game_number in enumerate(game_number_list):
        stat_ws.cell(row=row_idx+2, column=3, value=game_number)
    for row_idx, total_score in enumerate(total_score_list):
        stat_ws.cell(row=row_idx+2, column=4, value=total_score)



    report_name = rf'../report/{date.today()}-{datetime.now().strftime("%H%M%S%f")}.xlsx'
    stat_wb.save(report_name)
    print(f'Report saved at {report_name}')


for data_file in glob.glob(r'../data/*.xlsx'):
    verify_tournament_file_integrity(data_file)
update_player_stats(skip_keyword='incomplete')
